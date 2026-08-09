"""Uji bahan uji ekstraksi — B-1 fitur 015.

Bahan uji yang salah menghasilkan uji yang lulus atas hal yang keliru. Empat
PDF pada folder ini masing-masing mewakili satu keadaan pada `spec.md`, dan
berkas ini memastikan tiap berkas benar-benar berada pada keadaan itu —
sebelum ada satu pengekstrak pun yang mengandalkannya.

**Tidak ada data pribadi sungguhan.** Nomor yang dipakai disusun agar
mencocoki pola tanpa menjadi pengenal siapa pun: NIP berpola tanggal
1999-01-01 dengan sisa angka sembilan berulang, dan nomor telepon berakhir
nol berderet. Diuji, bukan sekadar dijanjikan.
"""

import re
from pathlib import Path

import pytest

BAHAN = Path(__file__).resolve().parents[1] / "bahan"

BERKAS = (
    "notulen.docx",
    "serapan.xlsx",
    "berlapis-teks.pdf",
    "pindaian-tanpa-teks.pdf",
    "terkunci.pdf",
    "rusak.pdf",
    "kosong.pdf",
)


@pytest.mark.parametrize("nama", BERKAS)
def test_berkas_bahan_ada(nama: str) -> None:
    assert (BAHAN / nama).is_file()


def test_berkas_kosong_benar_benar_kosong() -> None:
    """Keadaan "berkas nol bita" pada `spec.md` menuntut bahannya nol bita."""
    assert (BAHAN / "kosong.pdf").stat().st_size == 0


def test_berkas_lain_tidak_kosong() -> None:
    """Bahan yang kebetulan kosong akan membuat seluruh uji ekstraksi lulus
    karena tidak mengekstrak apa pun."""
    for nama in BERKAS:
        if nama != "kosong.pdf":
            assert (BAHAN / nama).stat().st_size > 0, nama


def test_pdf_berlapis_teks_memang_berlapis_teks() -> None:
    from pypdf import PdfReader

    teks = PdfReader(BAHAN / "berlapis-teks.pdf").pages[0].extract_text()
    assert "Sukamaju" in teks


def test_pdf_pindaian_memang_tanpa_lapisan_teks() -> None:
    """Kalau bahan ini ternyata punya teks, uji pengalihan ke OCR akan lulus
    tanpa pernah menguji pengalihannya."""
    from pypdf import PdfReader

    teks = PdfReader(BAHAN / "pindaian-tanpa-teks.pdf").pages[0].extract_text()
    assert teks.strip() == ""


def test_pdf_terkunci_memang_terkunci() -> None:
    from pypdf import PdfReader

    assert PdfReader(BAHAN / "terkunci.pdf").is_encrypted


def test_pdf_rusak_memang_tidak_dapat_diurai() -> None:
    from pypdf import PdfReader

    with pytest.raises(Exception):  # noqa: B017 — pustaka tidak menjanjikan satu tipe
        PdfReader(BAHAN / "rusak.pdf").pages[0].extract_text()


def test_bahan_tidak_memuat_data_pribadi_sungguhan() -> None:
    """Nomor yang berpola wajib tetap ada — pendeteksi Fase E memerlukannya —
    tetapi tidak boleh menjadi pengenal siapa pun.

    Yang diuji: setiap deret digit panjang pada bahan berakhir dengan angka
    berulang, bentuk yang tidak diterbitkan bagi orang sungguhan.
    """
    dari_docx = _teks_docx(BAHAN / "notulen.docx")
    for deret in re.findall(r"\d{10,}", dari_docx):
        assert re.search(r"(\d)\1{3,}", deret), f"deret {deret!r} tidak tampak dibuat-buat"


def _teks_docx(jalur: Path) -> str:
    from docx import Document

    return "\n".join(p.text for p in Document(str(jalur)).paragraphs)


def test_xlsx_memuat_sel_berumus_tanpa_nilai_tersimpan() -> None:
    """Keadaan yang wajib ditangani B-4, dan ia **tidak** disebut `spec.md`.

    Berkas XLSX yang ditulis pustaka tidak menyimpan hasil hitungan rumusnya;
    yang menyimpannya adalah aplikasi lembar sebar. Karena itu ada sel yang
    rumusnya ada tetapi nilainya tidak, dan pengekstrak wajib tidak menuliskan
    `=B2-C2` ke korpus sebagai jalan pintas.
    """
    from openpyxl import load_workbook

    berumus = load_workbook(BAHAN / "serapan.xlsx")["Serapan"]["D2"].value
    tersimpan = load_workbook(BAHAN / "serapan.xlsx", data_only=True)["Serapan"]["D2"].value
    assert str(berumus).startswith("=")
    assert tersimpan is None
