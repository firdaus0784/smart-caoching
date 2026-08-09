"""Pengekstrak XLSX — R-01, R-02, keadaan "XLSX berumus" pada spec.

**Nilai terhitung yang diambil, bukan rumusnya.** `=B2-C2` pada korpus bukan
sekadar sampah: ia rangkaian karakter yang akan diindeks, dicari, dan mungkin
dikutip sebagai bukti pada jawaban. Sebuah klaim manajerial yang bersitasi
pada potongan rumus adalah klaim bersitasi yang tidak berarti apa-apa, dan
validator sitasi fitur 008 tidak akan membedakannya.

**Keadaan ketiga yang tidak disebut `spec.md`: rumus ada, nilainya tidak.**
Format XLSX tidak mewajibkan hasil hitungan disimpan; yang menyimpannya adalah
aplikasi lembar sebar saat berkas terakhir dibuka. Berkas yang ditulis pustaka
atau dihasilkan sistem lain memuat rumus tanpa nilai tersimpan sama sekali.
Ditemukan saat menyusun bahan uji, bukan saat merancang.

Menulis rumusnya adalah jalan pintas yang paling menggoda pada keadaan itu,
karena rumusnya **selalu** tersedia ketika nilainya tidak. Yang dilakukan di
sini sebaliknya: sel dilewati, dan jumlah sel yang dilewati **dihitung serta
dibawa keluar** pada `sel_tak_terhitung`.

Menghitungnya bukan kerapian. Sel yang dilewati diam-diam menghasilkan laporan
anggaran yang kehilangan kolom sisa tanpa seorang pun tahu, dan verifikator
membaca tabel yang tampak lengkap. Angka itu yang memberinya alasan membuka
berkas aslinya.

Nama lembar ikut dibawa: satu berkas dapat memuat beberapa lembar dengan judul
kolom yang sama, dan tanpa namanya "Honor guru" pada lembar anggaran dan pada
lembar realisasi menjadi dua baris yang tidak dapat dibedakan siapa pun.
**Hanya bila lembarnya berisi**, karena nama lembar bukan isi — versi pertama
modul ini menuliskannya tanpa syarat, sehingga berkas yang seluruh selnya
kosong lolos sebagai dokumen bermuatan kata "Kosong".
"""

from __future__ import annotations

from pathlib import Path

from src.ingest.ekstraksi.dasar import Pengekstrak, TeksKanonik
from src.ingest.ekstraksi.galat import GalatEkstraksi

NAMA = "xlsx"


class TeksLembarSebar(TeksKanonik):
    """`TeksKanonik` dengan satu keterangan tambahan yang khas lembar sebar.

    Bukan bidang hiasan: ia satu-satunya tanda bahwa teks ini **tidak lengkap**
    terhadap berkas asalnya, dan ketidaklengkapan itu tidak terlihat dari
    isinya.
    """

    sel_tak_terhitung: int = 0


class PengekstrakXlsx(Pengekstrak):
    """Satu berkas XLSX menjadi satu `TeksLembarSebar`, atau `GalatEkstraksi`."""

    def menangani(self, jalur: Path) -> bool:
        return jalur.suffix.lower() == ".xlsx"

    def ekstrak(self, jalur: Path) -> TeksLembarSebar:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        try:
            berumus = load_workbook(jalur, data_only=False, read_only=True)
            bernilai = load_workbook(jalur, data_only=True, read_only=True)
        except (InvalidFileException, OSError, ValueError, KeyError) as galat:
            raise GalatEkstraksi(
                f"berkas XLSX tidak dapat dibuka: {type(galat).__name__}"
            ) from galat

        baris_teks: list[str] = []
        tak_terhitung = 0
        for nama_lembar in bernilai.sheetnames:
            baris_lembar: list[str] = []
            lembar_nilai = bernilai[nama_lembar]
            lembar_rumus = berumus[nama_lembar]
            for baris_nilai, baris_rumus in zip(
                lembar_nilai.iter_rows(values_only=True),
                lembar_rumus.iter_rows(values_only=True),
                strict=False,
            ):
                sel_terpakai: list[str] = []
                for nilai, rumus in zip(baris_nilai, baris_rumus, strict=False):
                    if nilai is not None:
                        sel_terpakai.append(str(nilai))
                    elif isinstance(rumus, str) and rumus.startswith("="):
                        tak_terhitung += 1
                if sel_terpakai:
                    baris_lembar.append("\t".join(sel_terpakai))
            # Nama lembar hanya ikut bila lembarnya berisi. Menuliskannya
            # tanpa syarat membuat berkas yang seluruh selnya kosong lolos
            # sebagai dokumen yang "berisi" nama lembarnya — cacat yang ada
            # pada versi pertama modul ini dan tertangkap ujinya.
            if baris_lembar:
                baris_teks.append(nama_lembar)
                baris_teks.extend(baris_lembar)

        isi = "\n".join(baris_teks)
        if not isi.strip():
            raise GalatEkstraksi("berkas XLSX terbuka tetapi tidak memuat isi")

        return TeksLembarSebar(
            isi=isi,
            asal=jalur.name,
            pengekstrak=NAMA,
            sel_tak_terhitung=tak_terhitung,
        )
